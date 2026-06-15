import urllib3
import openpyxl
from os import path
from argparse import ArgumentParser
from datetime import datetime
from catalystwan.session import create_manager_session
import json
from pygments import highlight, lexers, formatters

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


def tprint(message):
    """Print with timestamp prefix in format [HH:MM:SS DDMMYYYY]"""
    timestamp = datetime.now().strftime("[%H:%M:%S %d%m%Y]")
    print(f"{timestamp} {message}")


def pretty_print_dict_as_json(data):
    """Print dictionary as colored JSON format."""

    formatted_json = json.dumps(data, indent=4)
    colorful_json = highlight(
        formatted_json, lexers.JsonLexer(), formatters.TerminalFormatter()
    )

    print(colorful_json)


def get_static_vmanage_credentials():

    # Use this variables to hardcode credentials - ONLY FOR TESTING PURPOSES
    MANAGER_IP = ""
    USERNAME = ""
    PASSWORD = ""
    PORT = ""

    parser = ArgumentParser(description="Obtain vManage credentials from CLI arguments")
    parser.add_argument("--VMANAGE_IP", required=False, default=MANAGER_IP, help="VMANAGE IP or Hostname")
    parser.add_argument("--VMANAGE_USER", required=False, default=USERNAME, help="VMANAGE Username")
    parser.add_argument("--VMANAGE_PASSWORD", required=False, default=PASSWORD, help="VMANAGE Password")
    parser.add_argument("--VMANAGE_PORT", default=PORT if PORT else 443, help="VMANAGE Port", type=int)

    args = parser.parse_args()

    return (
        args.VMANAGE_IP,
        args.VMANAGE_USER,
        args.VMANAGE_PASSWORD,
        args.VMANAGE_PORT,
    )


def get_vmanage_credentials():

    parser = ArgumentParser(description="Obtain vManage credentials from CLI arguments")
    parser.add_argument("--VMANAGE_IP", required=True, help="VMANAGE IP or Hostname")
    parser.add_argument("--VMANAGE_USER", required=True, help="VMANAGE Username")
    parser.add_argument("--VMANAGE_PASSWORD", required=True, help="VMANAGE Password")
    parser.add_argument("--VMANAGE_PORT", default="443", help="VMANAGE Port", type=int)

    args = parser.parse_args()

    return (
        args.VMANAGE_IP,
        args.VMANAGE_USER,
        args.VMANAGE_PASSWORD,
        args.VMANAGE_PORT,
    )


def parse_data_from_get_request(session, endpoint):

    tprint("Fetching data from {}".format(endpoint))

    response = session.get(endpoint)
    data = response.json()["data"]

    return data


def get_device_config_data(session):

    endpoint = "dataservice/system/device/vedges"

    try:
        response = session.get(endpoint)
        data = response.json()["data"]

    except Exception as error:
        tprint(f"Failed to collect device config data: {error}")
        data = []

    return data


def sanitize_sheet_name(name):
    """Sanitize a string to be a valid Excel sheet name (max 31 chars, no special chars)."""
    invalid_chars = r"\/?*[]:"
    for char in invalid_chars:
        name = name.replace(char, "_")
    return name[:31]


def add_sheet(wb, sheet_name, data):
    """Add a new sheet to the workbook and write data rows into it."""
    if not data:
        tprint(f"No data to write for sheet '{sheet_name}'")
        return

    safe_name = sanitize_sheet_name(sheet_name)

    # Avoid duplicate sheet names by appending a counter
    existing_names = [ws.title for ws in wb.worksheets]
    if safe_name in existing_names:
        counter = 1
        while f"{safe_name[:28]}_{counter}" in existing_names:
            counter += 1
        safe_name = f"{safe_name[:28]}_{counter}"

    ws = wb.create_sheet(title=safe_name)

    headers = list(data[0].keys())
    ws.append(headers)

    for item in data:
        ws.append([str(v) if v is not None else "" for v in item.values()])

    tprint(f"Sheet '{safe_name}' written ({len(data)} rows)")


def add_summary_sheet(wb, template_info, device_info):
    """Add a formatted Summary sheet as the first sheet in the workbook."""
    ws = wb.create_sheet(title="Summary", index=0)

    # Styles
    title_font     = openpyxl.styles.Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    header_font    = openpyxl.styles.Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    cell_font      = openpyxl.styles.Font(name="Calibri", size=11)
    title_fill     = openpyxl.styles.PatternFill("solid", fgColor="1F4E79")
    header_fill    = openpyxl.styles.PatternFill("solid", fgColor="2E75B6")
    alt_fill       = openpyxl.styles.PatternFill("solid", fgColor="D6E4F0")
    center         = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    left           = openpyxl.styles.Alignment(horizontal="left",   vertical="center")
    thin_side      = openpyxl.styles.Side(style="thin", color="BBBBBB")
    border         = openpyxl.styles.Border(left=thin_side, right=thin_side,
                                            top=thin_side,  bottom=thin_side)

    # --- Title row ---
    ws.merge_cells("A1:C1")
    title_cell = ws["A1"]
    title_cell.value     = "Template Values — Summary"
    title_cell.font      = title_font
    title_cell.fill      = title_fill
    title_cell.alignment = center
    ws.row_dimensions[1].height = 28

    # --- Totals row ---
    ws.merge_cells("A2:C2")
    totals_cell = ws["A2"]
    totals_cell.value     = (
        f"Total templates: {len(template_info)}     "
        f"Total devices: {len(device_info)}"
    )
    totals_cell.font      = openpyxl.styles.Font(name="Calibri", bold=True, size=11, color="1F4E79")
    totals_cell.alignment = center
    ws.row_dimensions[2].height = 20

    # --- Blank spacer ---
    ws.row_dimensions[3].height = 8

    # --- Column headers ---
    headers = ["Template Name", "Template ID", "Attached Devices"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border
    ws.row_dimensions[4].height = 18

    # --- Data rows ---
    for row_idx, template in enumerate(template_info, start=5):
        fill = alt_fill if row_idx % 2 == 0 else openpyxl.styles.PatternFill()
        values = [
            template.get("template", ""),
            template.get("templateId", ""),
            len(template.get("associatedDevices", [])),
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font      = cell_font
            cell.fill      = fill
            cell.alignment = center if col == 3 else left
            cell.border    = border

    # --- Column widths ---
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 18

    tprint(f"Summary sheet written ({len(template_info)} templates)")


def save_workbook(wb):
    """Save the workbook to the script's root directory."""
    script_dir = path.dirname(path.abspath(__file__))
    filename = path.join(script_dir, "template_values.xlsx")
    wb.save(filename)
    tprint(f"Workbook saved: {filename}")


def filter_device_info(data):

    filter_by = "deviceType"
    filter_value = "vedge"
    interesting_keys = ["system-ip", "host-name", "site-id", "deviceModel", "reachability", "uuid", "template", "templateId"]
    device_info = []

    for i in data:
        if i.get(filter_by) == filter_value and i.get("system-ip"):
            device_info.append({key: i.get(key) for key in interesting_keys})

    return device_info


def get_template_info(device_info):

    template_info = []
    seen = set()

    for device in device_info:
        template_name = device.get("template")
        template_id = device.get("templateId")
        template_key = (template_name, template_id)

        if template_key not in seen and template_id:
            seen.add(template_key)
            template_info.append(
                {
                    "template": template_name,
                    "templateId": template_id,
                }
            )

    for template in template_info:
        associated_devices = []
        for device in device_info:
            if device.get("templateId") == template.get("templateId"):
                associated_devices.append(device.get("uuid"))
        template["associatedDevices"] = associated_devices

    return template_info


def get_values_per_template(session, template):

    endpoint = "/dataservice/template/device/config/input/"
    template_values = []

    payload = {
        "templateId": template.get("templateId"),
        "deviceIds": template.get("associatedDevices"),
    }

    response = session.post(url=endpoint, json=payload)
    device_values = response.json()["data"]
    template_values.extend(device_values)

    return template_values


if __name__ == "__main__":

    url, username, password, port = get_static_vmanage_credentials()
    if not all([url, username, password, port]):
        url, username, password, port = get_vmanage_credentials()

    try:
        with create_manager_session(
            url=url,
            username=username,
            password=password,
            port=int(port),
        ) as session:

            if session:
                tprint("vManage Session created successfully!")

                wb = openpyxl.Workbook()
                # Remove the default empty sheet created by openpyxl
                wb.remove(wb.active)

                data = get_device_config_data(session)
                device_info = filter_device_info(data)
                tprint("Device info collected for {} devices".format(len(device_info)))

                template_info = get_template_info(device_info)
                tprint("Template info collected for {} templates".format(len(template_info)))

                add_summary_sheet(wb, template_info, device_info)

                for template in template_info:
                    template_values = get_values_per_template(session, template)
                    add_sheet(wb, template.get("template", "unknown"), template_values)

                save_workbook(wb)

            else:
                tprint("Failed to create vManage session. Please check your credentials and try again.")

    except Exception as error:
        tprint(f"An error occurred: {error}")
